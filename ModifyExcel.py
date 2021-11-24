from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

property_file = "202111.xlsx"
#請使用者輸入資料檔案
def user_input():
	filename = input('請輸入要處理的Excel檔案(打q結束程式): ')
	if filename == 'q':
		filename = ''    #將檔名清空
		print('感謝您的使用!')
	return filename

#讀取付款條件參數
def read_parmeter1(filename):
	wb = load_workbook('files/' + filename)
	for s in range(len(wb.sheetnames)):
		if wb.sheetnames[s] == 'S1_參數':
			break
	wb.active = s
	ws = wb.active

	parameter = {}
	for row in range(4, 34):
		name = ''
		code = ''
		for col in range(8, 10):
			char = get_column_letter(col)
			if col == 8:
				name = ws[char + str(row)].value
				#item.append(name)
			elif col == 9:
				code = ws[char + str(row)].value
		parameter[code] = name

	#加入特殊條件
	parameter['放後10天'] = parameter['放10天']
	parameter['訂後10天'] = parameter['訂10天']
	print(parameter)
	return parameter

#讀取接單人員參數
def read_parmeter2(filename):
	wb = load_workbook('files/' + filename)
	for s in range(len(wb.sheetnames)):
		if wb.sheetnames[s] == 'S1_參數':
			break
	wb.active = s
	ws = wb.active

	parameter = {}
	for row in range(4, 24):
		name = ''
		code = 0
		for col in range(5, 7):
			char = get_column_letter(col)
			if col == 5:
				name = int(ws[char + str(row)].value)
				#item.append(name)
			elif col == 6:
				code = ws[char + str(row)].value
		parameter[code] = name
	#加入特殊條件
	parameter['王'] = parameter['王琦蒲']
	parameter['蒲'] = parameter['王琦蒲']
	parameter['琦'] = parameter['王琦蒲']
	parameter['彥'] = parameter['蔣聖彥']
	parameter['蔣'] = parameter['蔣聖彥']
	parameter['宗'] = parameter['呂宗憲']
	parameter['憲'] = parameter['呂宗憲']
	parameter['信'] = parameter['陳永信']
	parameter['怡'] = parameter['潘俊怡']
	parameter['潘'] = parameter['潘俊怡']
	parameter['勳'] = parameter['魏孝勳']
	parameter['孝'] = parameter['魏孝勳']
	parameter['梅'] = parameter['廖麗梅']
	parameter['C'] = parameter['曹婉棋']
	parameter['文'] = parameter['薛舒文']
	parameter['欣'] = parameter['林吟欣']
	parameter['吟'] = parameter['林吟欣']
	parameter['棋'] = parameter['曹婉棋']
	parameter['永'] = parameter['陳永信']

	print(parameter)
	return parameter

#主要進行轉換
def proceser(filename, payment_type, sales_person):
	# 進行[M2_已建未收PO]的轉換
	print('進行PO單轉換.........')
	wb = load_workbook('files/' + filename)
	for s in range(len(wb.sheetnames)):
		if wb.sheetnames[s] == 'M2_已建未收PO':
			break
	wb.active = s
	ws = wb.active
	for row in range(9, 150):
		char = get_column_letter(17)
		key = str(ws[char + str(row)].value)
		if key in payment_type:
			ws[char + str(row)].value = payment_type[key]
			print(str(row) + '- Key:' + key +'--> 轉換為:' + payment_type[key])
		else:
			print(str(row) + '- Key:' + key + '--> No matched, Skip this row')

	# 進行[S2_已建單未出貨的SO]的轉換
	print('進行SO單轉換.........')
	for s in range(len(wb.sheetnames)):
		if wb.sheetnames[s] == 'S2_已建單未出貨的SO':
			break
	wb.active = s
	ws = wb.active
	for row in range(9, 587):
		char1 = get_column_letter(16)
		char2 = get_column_letter(19)
		key1 = str(ws[char1 + str(row)].value)
		key2 = str(ws[char2 + str(row)].value)
		if key1 in payment_type:
			ws[char1 + str(row)].value = payment_type[key1]
			print(str(row) + '- Key:' + key1 +'--> 轉換為:' + payment_type[key1])
		else:
			print(str(row) + '- Key:' + key1 + '--> No matched, Skip this row')

		if key2 in sales_person:
			ws[char2 + str(row)].value = str(sales_person[key2])
			print(str(row) + '- Key:' + key2 +'--> 轉換為:' + str(sales_person[key2]))
		elif key2 != 'None':
			ws[char2 + str(row)].value = str(sales_person['其它'])
			print(str(row) + '- Key:' + key2 + '長度:' + str(len(key2)) + ' --> No matched, 轉換為' + str(sales_person['其它']))
		else:
			print(str(row) + '- Key:' + key2 + "No data, skip掉")

	wb.save('files/' + filename)


#主程式區塊
def main():
	filename = user_input()
	if filename != '':
		payment_type = read_parmeter1('202111.xlsx')
		sales_person = read_parmeter2('202111.xlsx')
		proceser(filename, payment_type, sales_person)

#執行
main()
